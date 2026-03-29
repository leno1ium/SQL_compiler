from pathlib import Path
from openpyxl.utils import datetime
from SQL_compiler.executor.table import ExcelLoader
from SQL_compiler.executor.executor import QueryExecutor
from SQL_compiler.parser.parser import parse


def load_excel_file() -> Path:

    while True:
        print("\nВыберите действие:")
        print("1. Указать путь к существующему Excel файлу")
        print("2. Создать тестовый файл data.xlsx")
        print("3. Выйти")

        choice = input("\nВаш выбор (1/2/3): ").strip()

        if choice == "1":
            file_path = input("Введите путь к Excel файлу: ").strip()
            path = Path(file_path)

            if not path.exists():
                print(f"Файл не найден: {path}")
                continue

            if not path.suffix.lower() in ['.xlsx', '.xls']:
                print(f"Файл должен быть Excel (.xlsx или .xls)")
                continue

            return path

        elif choice == "2":
            return create_sample_excel()

        elif choice == "3":
            print("Выход из программы...")
            exit(0)

        else:
            print("Неверный выбор. Попробуйте снова.")


def create_sample_excel() -> Path:
    import openpyxl

    print("\nСоздание тестового Excel файла...")

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "users"

    headers = ["id", "name", "age", "email", "city"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    users_data = [
        [1, "Alice Smith", 25, "alice@mail.com", "Moscow"],
        [2, "Bob Johnson", 30, "bob@mail.com", "SPB"],
        [3, "Charlie Brown", 22, "charlie@mail.com", "Moscow"],
        [4, "Diana Ross", 28, "diana@mail.com", "Kazan"],
        [5, "Eve Wilson", 35, None, "Moscow"],
        [6, "Frank Miller", 19, "frank@mail.com", "SPB"],
        [7, "Grace Lee", 42, "grace@mail.com", "Moscow"],
        [8, "Henry Ford", 55, "henry@mail.com", "SPB"],
        [9, "Ivy Chen", 24, None, "Kazan"],
        [10, "Jack Smith", 31, "jack@mail.com", "Moscow"],
        [11, "Kevin Brown", 27, "kevin@mail.com", "SPB"],
        [12, "Laura Wilson", 29, "laura@mail.com", "Moscow"],
    ]

    for row_idx, row in enumerate(users_data, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws2 = wb.create_sheet("orders")

    headers2 = ["id", "user_id", "product_id", "total", "order_date"]
    for col, header in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=header)

    orders_data = [
        [1, 1, 1, 1500.50, "2023-01-15"],
        [2, 2, 2, 2500.00, "2023-01-20"],
        [3, 1, 3, 800.75, "2023-02-01"],
        [4, 3, 1, 1500.50, "2023-02-10"],
        [5, 4, 2, 2500.00, "2023-02-15"],
        [6, 1, 4, 1200.00, "2023-03-01"],
        [7, 5, 1, 1500.50, "2023-03-05"],
        [8, 2, 3, 800.75, "2023-03-10"],
        [9, 6, 4, 1200.00, "2023-03-15"],
        [10, 7, 2, 2500.00, "2023-03-20"],
    ]

    for row_idx, row in enumerate(orders_data, 2):
        for col_idx, value in enumerate(row, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)

    ws3 = wb.create_sheet("products")

    headers3 = ["id", "name", "price", "category"]
    for col, header in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=header)

    products_data = [
        [1, "Laptop", 1500.50, "Electronics"],
        [2, "Smartphone", 2500.00, "Electronics"],
        [3, "Headphones", 800.75, "Electronics"],
        [4, "Mouse", 1200.00, "Accessories"],
        [5, "Keyboard", 1800.25, "Accessories"],
        [6, "Monitor", 3200.00, "Electronics"],
        [7, "Tablet", 2000.00, "Electronics"],
        [8, "USB Cable", 500.00, "Accessories"],
    ]

    for row_idx, row in enumerate(products_data, 2):
        for col_idx, value in enumerate(row, 1):
            ws3.cell(row=row_idx, column=col_idx, value=value)

    ws4 = wb.create_sheet("sales")

    headers4 = ["id", "product_name", "category", "amount", "sale_date"]
    for col, header in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=header)

    sales_data = [
        [1, "Laptop", "Electronics", 2, "2023-01-15"],
        [2, "Mouse", "Accessories", 5, "2023-01-16"],
        [3, "Smartphone", "Electronics", 1, "2023-01-17"],
        [4, "Keyboard", "Accessories", 3, "2023-01-18"],
        [5, "Laptop", "Electronics", 1, "2023-01-19"],
        [6, "Headphones", "Electronics", 4, "2023-01-20"],
        [7, "Mouse", "Accessories", 2, "2023-01-21"],
        [8, "Monitor", "Electronics", 1, "2023-01-22"],
        [9, "Laptop", "Electronics", 3, "2023-01-23"],
        [10, "Keyboard", "Accessories", 2, "2023-01-24"],
        [11, "Smartphone", "Electronics", 2, "2023-01-25"],
        [12, "Mouse", "Accessories", 4, "2023-01-26"],
    ]

    for row_idx, row in enumerate(sales_data, 2):
        for col_idx, value in enumerate(row, 1):
            ws4.cell(row=row_idx, column=col_idx, value=value)

    filepath = Path("data.xlsx")
    wb.save(filepath)

    print(f"Создан тестовый Excel файл: {filepath.absolute()}")

    return filepath


def print_full_table(table, title: str = None):
    if title:
        print(f"\n{title}")
        print("=" * len(title))

    print("\nСХЕМА ТАБЛИЦЫ:")
    print("-" * 60)
    print(f"Имя таблицы: {table.name}")
    print(f"Всего колонок: {len(table.column_names)}")
    print(f"Всего строк: {len(table.rows)}")

    print("\nКОЛОНКИ И ТИПЫ:")
    print("-" * 60)
    print(f"{'№':<5} {'Колонка':<20} {'Тип':<15} {'Примеры значений':<30}")
    print("-" * 60)

    for i, col in enumerate(table.column_names, 1):
        col_type = table.column_types[col]
        type_name = {
            int: "INTEGER",
            float: "FLOAT",
            str: "STRING",
            datetime: "DATE"
        }.get(col_type, str(col_type))

        examples = []
        for j in range(min(3, len(table.rows))):
            val = table.rows[j].get(col)
            if val is not None:
                if isinstance(val, float):
                    examples.append(f"{val:.2f}")
                else:
                    examples.append(str(val))

        example_str = ", ".join(examples) if examples else "(нет данных)"
        print(f"{i:<5} {col:<20} {type_name:<15} {example_str:<30}")

    print("\nВСЕ ДАННЫЕ ТАБЛИЦЫ:")
    print("-" * 100)

    header_line = " | ".join(f"{col:<15}" for col in table.column_names)
    print(header_line)
    print("-" * 100)

    for row_idx, row in enumerate(table.rows, 1):
        row_values = []
        for col in table.column_names:
            val = row.get(col)
            if val is None:
                row_values.append(f"{'NULL':<15}")
            elif isinstance(val, float):
                row_values.append(f"{val:<15.2f}")
            else:
                row_values.append(f"{str(val):<15}")

        print(f"{row_idx:<3} " + " | ".join(row_values))

    print("-" * 100)
    print(f"Всего выведено строк: {len(table.rows)}")


def run_test(executor: QueryExecutor, sql: str, description: str):
    print("\n" + "=" * 80)
    print(f"{description}")
    print(f"SQL: {sql}")
    print("=" * 80)

    try:
        ast = parse(sql)

        print("\nAST ДЕРЕВО:")
        for line in ast.tree:
            print(line)

        result = executor.execute(ast)

        if result:
            headers = list(result[0].keys())

            header_line = " | ".join(f"{h:<15}" for h in headers)
            print(header_line)
            print("-" * 80)

            for row in result:
                row_values = []
                for v in row.values():
                    if v is None:
                        row_values.append(f"{'NULL':<15}")
                    elif isinstance(v, float):
                        row_values.append(f"{v:<15.2f}")
                    else:
                        row_values.append(f"{str(v):<15}")
                print(" | ".join(row_values))

            if len(result) > 20:
                print(f"... и еще {len(result) - 20} строк")
        else:
            print("Нет данных, удовлетворяющих условию")

    except Exception as e:
        print(f"Ошибка: {e}")

    input("\nНажмите Enter для продолжения...")


def verify_test_data(tables):
    print("\nПРОВЕРКА НАЛИЧИЯ ТАБЛИЦ:")
    print("=" * 60)

    required_tables = ["users", "orders", "products", "sales"]
    missing_tables = []

    for table_name in required_tables:
        if table_name in tables:
            table = tables[table_name]
            print(f"Таблица '{table_name}' найдена ({len(table)} строк)")

            if table_name == "users":
                a_names = [row for row in table.rows if row.get('name', '').startswith('A')]
                print(f"   • Имена на 'A': {len(a_names)} шт. - {[row['name'] for row in a_names]}")

                null_emails = [row for row in table.rows if row.get('email') is None]
                print(f"   • NULL email: {len(null_emails)} шт.")

            elif table_name == "orders":
                user_ids = set(row['user_id'] for row in table.rows)
                print(f"   • user_id в заказах: {sorted(user_ids)}")

        else:
            print(f"Таблица '{table_name}' НЕ найдена")
            missing_tables.append(table_name)

    if missing_tables:
        print(f"\nОтсутствуют таблицы: {', '.join(missing_tables)}")
        print("   Тесты, обращающиеся к этим таблицам, будут падать с ошибкой 'Table not found'")
    else:
        print(f"\nВсе необходимые таблицы присутствуют!")

    print("=" * 60)


def load_tests_from_file() -> list:
    tests = []

    try:
        from SQL_compiler.test import test as test_module

        if hasattr(test_module, 'tests'):
            tests_from_file = test_module.tests

            for i, sql in enumerate(tests_from_file, 1):
                clean_sql = sql.strip().rstrip(';')
                description = f"Тест {i}: {clean_sql[:50]}..." if len(clean_sql) > 50 else clean_sql
                tests.append((sql, description))
        else:
            print("\nВ файле test.py нет списка 'tests'")

    except ImportError:
        print("\nФайл tests.py не найден. Используются встроенные тесты.")
        builtin_tests = [
            ("SELECT * FROM students", "Базовый SELECT *"),
            ("SELECT name, age FROM students WHERE age > 20", "WHERE условие"),
            ("SELECT name FROM students WHERE name LIKE 'И%'", "LIKE оператор"),
        ]
        return builtin_tests

    except Exception as e:
        print(f"\nОшибка при загрузке test.py: {e}")
        return []

    return tests


def main():

    excel_file = load_excel_file()
    loader = ExcelLoader()

    try:
        tables = loader.load(excel_file)

        for name, table in tables.items():
            print_full_table(table, f"ТАБЛИЦА: {name}")

    except Exception as e:
        print(f"Ошибка загрузки Excel: {e}")
        return

    executor = QueryExecutor(tables)

    tests = load_tests_from_file()

    if not tests:
        print("\nНет тестов для выполнения. Добавьте тесты в tests.py")
        return

    for sql, description in tests:
        run_test(executor, sql, description)


if __name__ == "__main__":
    main()