from typing import List, Dict, Any, Optional, Union
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


class Table:

    INTEGER = int
    FLOAT = float
    STRING = str
    DATE = datetime

    def __init__(self, name: str):
        self.name = name
        self.column_names: List[str] = []
        self.column_types: Dict[str, type] = {}
        self.rows: List[Dict[str, Any]] = []

    def add_column(self, name: str, col_type: type) -> None:
        self.column_names.append(name)
        self.column_types[name] = col_type

    def add_row(self, row: Dict[str, Any]) -> None:
        for col in self.column_names:
            if col not in row:
                row[col] = None
        self.rows.append(row.copy())

    def get_value(self, row_idx: int, column: str) -> Any:
        if row_idx < 0 or row_idx >= len(self.rows):
            raise IndexError(f"Row index {row_idx} out of range (0-{len(self.rows) - 1})")
        if column not in self.column_names:
            raise KeyError(f"Column '{column}' not found in table '{self.name}'")

        return self.rows[row_idx].get(column)

    def get_column_type(self, column: str) -> type:
        if column not in self.column_types:
            raise KeyError(f"Column '{column}' not found in table '{self.name}'")
        return self.column_types[column]

    def get_column_index(self, column: str) -> int:
        if column not in self.column_names:
            raise KeyError(f"Column '{column}' not found in table '{self.name}'")
        return self.column_names.index(column)

    def __len__(self) -> int:
        return len(self.rows)

    def __str__(self) -> str:
        return f"Table('{self.name}', columns={len(self.column_names)}, rows={len(self.rows)})"

    def print_schema(self):
        print(f"\nТаблица: {self.name}")
        print("-" * 50)
        print(f"{'Колонка':<20} {'Тип':<15} {'Примеры':<30}")
        print("-" * 50)

        for col in self.column_names:
            col_type = self.column_types[col]
            type_name = {
                int: "INTEGER",
                float: "FLOAT",
                str: "STRING",
                datetime: "DATE"
            }.get(col_type, str(col_type))

            examples = []
            for i in range(min(3, len(self.rows))):
                val = self.rows[i].get(col)
                if val is not None:
                    if isinstance(val, datetime):
                        examples.append(val.strftime("%Y-%m-%d"))
                    else:
                        examples.append(str(val))

            example_str = ", ".join(examples) if examples else "(нет данных)"
            print(f"{col:<20} {type_name:<15} {example_str:<30}")

    def print_data(self) -> None:
        print(f"\nДанные таблицы {self.name}:")
        print("-" * 80)

        header = " | ".join(f"{col:<15}" for col in self.column_names)
        print(header)
        print("-" * 80)

        for i, row in enumerate(self.rows):
            row_str = " | ".join(
                f"{str(row.get(col, 'NULL')):<15}"
                for col in self.column_names
            )
            print(row_str)

        print(f"\nВсего строк: {len(self.rows)}")


class ExcelLoader:

    def __init__(self):
        self.tables: Dict[str, Table] = {}

    def load(self, filepath: Union[str, Path]) -> Dict[str, Table]:
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {filepath}")

        print(f"Загрузка Excel файла: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"  Обработка листа: {sheet_name}")
            table = self._load_sheet(sheet_name, sheet)
            self.tables[sheet_name] = table

        wb.close()
        print(f"Загружено таблиц: {len(self.tables)}")
        return self.tables

    def _load_sheet(self, name: str, sheet) -> Table:
        table = Table(name)

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            print(f"  Лист '{name}' пуст")
            return table

        headers = []
        for i, cell in enumerate(rows[0]):
            if cell is None:
                header = f"Column{i + 1}"
            else:
                header = str(cell).strip()
            headers.append(header)

        print(f"    Найдено колонок: {len(headers)}")

        data_rows = rows[1:] if len(rows) > 1 else []
        print(f"    Найдено строк данных: {len(data_rows)}")

        columns_data = [[] for _ in headers]
        for row in data_rows:
            for i, value in enumerate(row):
                if i < len(columns_data):
                    columns_data[i].append(value)

        for i, header in enumerate(headers):
            col_type = self._detect_type(columns_data[i])
            table.add_column(header, col_type)
            print(f"    Колонка '{header}': {col_type.__name__}")

        for row_idx, row in enumerate(data_rows):
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    value = row[i]
                    target_type = table.get_column_type(header)
                    row_dict[header] = self._convert_value(value, target_type)
                else:
                    row_dict[header] = None
            table.add_row(row_dict)

        return table

    def _detect_type(self, values: List[Any]) -> type:
        if not values:
            return str

        non_null = [v for v in values if v is not None]
        if not non_null:
            return str

        all_int = True
        all_float = True
        all_date = True

        for v in non_null:
            if all_int:
                if isinstance(v, int):
                    continue
                elif isinstance(v, float) and v.is_integer():
                    continue
                elif isinstance(v, str) and v.strip().lstrip('-').isdigit():
                    continue
                else:
                    all_int = False

            if all_float:
                if isinstance(v, (int, float)):
                    continue
                elif isinstance(v, str):
                    try:
                        float(v)
                        continue
                    except ValueError:
                        all_float = False
                else:
                    all_float = False

            if all_date:
                if isinstance(v, datetime):
                    continue
                elif isinstance(v, str):
                    date_formats = [
                        "%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d",
                        "%d-%m-%Y", "%Y%m%d", "%d.%m.%y"
                    ]
                    for fmt in date_formats:
                        try:
                            datetime.strptime(v, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        all_date = False
                else:
                    all_date = False

        if all_int:
            return int
        if all_float:
            return float
        if all_date:
            return datetime
        return str

    def _convert_value(self, value: Any, target_type: type) -> Any:
        if value is None:
            return None

        try:
            if target_type == int:
                if isinstance(value, float) and value.is_integer():
                    return int(value)
                if isinstance(value, str):
                    return int(value.strip())
                return int(value)

            elif target_type == float:
                return float(value)

            elif target_type == datetime:
                if isinstance(value, datetime):
                    return value
                if isinstance(value, str):
                    date_formats = [
                        "%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d",
                        "%d-%m-%Y", "%Y%m%d", "%d.%m.%y"
                    ]
                    for fmt in date_formats:
                        try:
                            return datetime.strptime(value, fmt)
                        except ValueError:
                            continue
                return None

            else:
                return str(value).strip()

        except (ValueError, TypeError):
            return None